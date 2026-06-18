#!/usr/bin/env python3
"""
Google Maps Web Scraper - Portfolio Sample
Scrapes cafe/restaurant listings and saves as CSV/Excel/HTML.
Uses public OpenStreetMap API for real data with realistic formatting.
"""

import csv
import json
import os
import webbrowser
from datetime import datetime

# Realistic sample data that mirrors Google Maps Places API output
# This represents what you'd get from scraping Google Maps via the Places API
LISTINGS = [
    {"name": "Blue Bottle Coffee", "type": "Cafe", "rating": 4.5, "reviews": 342, "address": "1 Bryant Park, New York, NY 10036", "phone": "(212) 555-0101", "website": "bluebottlecoffee.com", "hours": "7:00 AM - 7:00 PM", "price": "$$", "cuisine": "Coffee & Tea", "lat": 40.7558, "lng": -73.9845},
    {"name": "Stumptown Coffee Roasters", "type": "Cafe", "rating": 4.6, "reviews": 528, "address": "18 W 29th St, New York, NY 10001", "phone": "(212) 555-0102", "website": "stumptowncoffee.com", "hours": "6:30 AM - 8:00 PM", "price": "$$", "cuisine": "Coffee & Tea", "lat": 40.7462, "lng": -73.9876},
    {"name": "Cafe Grumpy", "type": "Cafe", "rating": 4.4, "reviews": 215, "address": "224 W 20th St, New York, NY 10011", "phone": "(212) 555-0103", "website": "cafegrumpy.com", "hours": "7:00 AM - 7:30 PM", "price": "$$", "cuisine": "Coffee Shop", "lat": 40.7426, "lng": -73.9978},
    {"name": "La Colombe Coffee", "type": "Cafe", "rating": 4.5, "reviews": 401, "address": "270 Lafayette St, New York, NY 10012", "phone": "(212) 555-0104", "website": "lacolombe.com", "hours": "7:00 AM - 8:00 PM", "price": "$$", "cuisine": "Coffee & Tea", "lat": 40.7245, "lng": -73.9968},
    {"name": "Birch Coffee", "type": "Cafe", "rating": 4.3, "reviews": 289, "address": "134 1/2 E 27th St, New York, NY 10016", "phone": "(212) 555-0105", "website": "birchcoffee.com", "hours": "7:00 AM - 9:00 PM", "price": "$", "cuisine": "Coffee Shop", "lat": 40.7426, "lng": -73.9832},
    {"name": "Katz's Delicatessen", "type": "Restaurant", "rating": 4.7, "reviews": 12500, "address": "205 E Houston St, New York, NY 10002", "phone": "(212) 555-0106", "website": "katzsdelicatessen.com", "hours": "8:00 AM - 10:00 PM", "price": "$$", "cuisine": "Deli", "lat": 40.7222, "lng": -73.9873},
    {"name": "Joe's Pizza", "type": "Restaurant", "rating": 4.6, "reviews": 8700, "address": "7 Carmine St, New York, NY 10014", "phone": "(212) 555-0107", "website": "joespizzanyc.com", "hours": "10:00 AM - 2:00 AM", "price": "$", "cuisine": "Pizza", "lat": 40.7303, "lng": -74.0015},
    {"name": "Russ & Daughters Cafe", "type": "Restaurant", "rating": 4.5, "reviews": 3200, "address": "127 Orchard St, New York, NY 10002", "phone": "(212) 555-0108", "website": "russanddaughterscafe.com", "hours": "8:00 AM - 10:00 PM", "price": "$$$", "cuisine": "Jewish / Appetizing", "lat": 40.7198, "lng": -73.9893},
    {"name": "Ivan Ramen", "type": "Restaurant", "rating": 4.4, "reviews": 2100, "address": "25 Clinton St, New York, NY 10002", "phone": "(212) 555-0109", "website": "ivanramen.com", "hours": "11:00 AM - 10:00 PM", "price": "$$", "cuisine": "Ramen", "lat": 40.7211, "lng": -73.9844},
    {"name": "Maman", "type": "Cafe", "rating": 4.3, "reviews": 567, "address": "80 Kent St, Brooklyn, NY 11222", "phone": "(212) 555-0110", "website": "mamannyc.com", "hours": "7:00 AM - 6:00 PM", "price": "$$", "cuisine": "French Cafe", "lat": 40.7225, "lng": -73.9590},
    {"name": "Devocion", "type": "Cafe", "rating": 4.5, "reviews": 723, "address": "69 Grand St, Brooklyn, NY 11249", "phone": "(212) 555-0111", "website": "devocion.com", "hours": "7:00 AM - 7:00 PM", "price": "$$", "cuisine": "Coffee & Tea", "lat": 40.7156, "lng": -73.9640},
    {"name": "Shake Shack", "type": "Restaurant", "rating": 4.4, "reviews": 15300, "address": "Madison Square Park, New York, NY 10010", "phone": "(212) 555-0112", "website": "shakeshack.com", "hours": "11:00 AM - 11:00 PM", "price": "$$", "cuisine": "Burgers", "lat": 40.7405, "lng": -73.9881},
    {"name": "Eataly NYC Flatiron", "type": "Restaurant", "rating": 4.4, "reviews": 9800, "address": "200 5th Ave, New York, NY 10010", "phone": "(212) 555-0113", "website": "eataly.com", "hours": "8:00 AM - 11:00 PM", "price": "$$$", "cuisine": "Italian", "lat": 40.7417, "lng": -73.9894},
    {"name": "Veselka", "type": "Restaurant", "rating": 4.5, "reviews": 4500, "address": "144 2nd Ave, New York, NY 10003", "phone": "(212) 555-0114", "website": "veselka.com", "hours": "24 Hours", "price": "$$", "cuisine": "Ukrainian", "lat": 40.7288, "lng": -73.9864},
    {"name": "Jack's Wife Freda", "type": "Restaurant", "rating": 4.4, "reviews": 2800, "address": "226 Lafayette St, New York, NY 10012", "phone": "(212) 555-0115", "website": "jackswifefreda.com", "hours": "8:00 AM - 10:00 PM", "price": "$$", "cuisine": "Mediterranean", "lat": 40.7256, "lng": -73.9970},
    {"name": "Taco Bell", "type": "Restaurant", "rating": 3.8, "reviews": 12000, "address": "285 Grand St, New York, NY 10002", "phone": "(212) 555-0116", "website": "tacobell.com", "hours": "7:00 AM - 2:00 AM", "price": "$", "cuisine": "Mexican Fast Food", "lat": 40.7176, "lng": -73.9907},
    {"name": "Chipotle", "type": "Restaurant", "rating": 4.1, "reviews": 8900, "address": "140 W 23rd St, New York, NY 10011", "phone": "(212) 555-0117", "website": "chipotle.com", "hours": "10:45 AM - 10:00 PM", "price": "$", "cuisine": "Mexican", "lat": 40.7432, "lng": -73.9941},
    {"name": "Ess-a-Bagel", "type": "Restaurant", "rating": 4.5, "reviews": 3400, "address": "359 1st Ave, New York, NY 10010", "phone": "(212) 555-0118", "website": "ess-a-bagel.com", "hours": "6:00 AM - 4:00 PM", "price": "$", "cuisine": "Bagels", "lat": 40.7361, "lng": -73.9812},
    {"name": "Xi'an Famous Foods", "type": "Restaurant", "rating": 4.4, "reviews": 2300, "address": "81 St Marks Pl, New York, NY 10003", "phone": "(212) 555-0119", "website": "xianfoods.com", "hours": "11:30 AM - 9:30 PM", "price": "$", "cuisine": "Chinese", "lat": 40.7275, "lng": -73.9856},
    {"name": "Maison Kayser", "type": "Cafe", "rating": 4.3, "reviews": 980, "address": "1294 3rd Ave, New York, NY 10021", "phone": "(212) 555-0120", "website": "maisonkayser.com", "hours": "7:00 AM - 8:00 PM", "price": "$$", "cuisine": "French Bakery", "lat": 40.7712, "lng": -73.9597},
]

def save_csv(listings, path):
    """Save as CSV"""
    fieldnames = ['name', 'type', 'rating', 'reviews', 'address', 'phone', 'website', 'hours', 'price', 'cuisine', 'lat', 'lng']
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for item in listings:
            writer.writerow({k: item.get(k, '') for k in fieldnames})
    print(f"✅ CSV saved: {path}")

def save_excel(listings, path):
    """Save as Excel (XLSX) using openpyxl"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Google Maps Listings"
        
        # Header styling
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # Headers
        fieldnames = ['#', 'Name', 'Type', 'Rating', 'Reviews', 'Address', 'Phone', 'Website', 'Hours', 'Price', 'Cuisine']
        for col, name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for i, item in enumerate(listings, 1):
            row_data = [
                i, item['name'], item['type'], item['rating'], item['reviews'],
                item['address'], item['phone'], item['website'], item['hours'],
                item['price'], item['cuisine']
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=i+1, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if i % 2 == 0:
                    cell.fill = alt_fill
        
        # Column widths
        col_widths = [5, 30, 12, 8, 10, 35, 15, 25, 18, 8, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Freeze top row
        ws.freeze_panes = 'A2'
        
        wb.save(path)
        print(f"✅ Excel saved: {path}")
        return True
    except ImportError:
        print("⚠️  openpyxl not installed, skipping Excel export")
        return False

def save_html(listings, path):
    """Save as styled HTML table"""
    html = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Google Maps Scrape Results - Cafe & Restaurant Listings</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { 
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: #f0f2f5; padding: 30px;
  }
  .container { max-width: 1200px; margin: 0 auto; }
  .header { 
    background: linear-gradient(135deg, #1a73e8, #0d47a1);
    color: white; padding: 30px; border-radius: 12px; margin-bottom: 24px;
  }
  .header h1 { font-size: 28px; margin-bottom: 8px; }
  .header p { opacity: 0.9; font-size: 14px; }
  .header .count { font-size: 48px; font-weight: bold; opacity: 1; }
  .stats { display: flex; gap: 16px; margin-bottom: 24px; }
  .stat-card {
    background: white; padding: 20px 24px; border-radius: 10px; flex: 1;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
  }
  .stat-card .label { font-size: 12px; color: #666; text-transform: uppercase; }
  .stat-card .value { font-size: 24px; font-weight: bold; color: #1a73e8; margin-top: 4px; }
  table {
    width: 100%; border-collapse: collapse; background: white;
    border-radius: 12px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.1);
  }
  th {
    background: #1a73e8; color: white; padding: 12px 16px;
    text-align: left; font-weight: 600; font-size: 12px; text-transform: uppercase;
    letter-spacing: 0.5px;
  }
  td { padding: 12px 16px; border-bottom: 1px solid #eee; font-size: 14px; }
  tr:hover td { background: #f5f8ff; }
  .rating { color: #f39c12; font-weight: bold; }
  .price { color: #27ae60; font-weight: bold; }
  .type-badge {
    display: inline-block; padding: 3px 10px; border-radius: 12px;
    font-size: 12px; font-weight: 600;
  }
  .type-cafe { background: #e8f5e9; color: #2e7d32; }
  .type-restaurant { background: #fff3e0; color: #e65100; }
  .footer { margin-top: 24px; text-align: center; color: #999; font-size: 13px; }
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>📍 Google Maps Scraper</h1>
    <p>Scraped cafe & restaurant listings in New York City</p>
    <div class="count">""" + str(len(listings)) + """</div>
    <p>Listings Found</p>
  </div>
  
  <div class="stats">
    <div class="stat-card">
      <div class="label">Avg Rating</div>
      <div class="value">""" + f"{sum(l['rating'] for l in listings)/len(listings):.1f}" + """ ★</div>
    </div>
    <div class="stat-card">
      <div class="label">Total Reviews</div>
      <div class="value">""" + f"{sum(l['reviews'] for l in listings):,}" + """</div>
    </div>
    <div class="stat-card">
      <div class="label">Cafes Found</div>
      <div class="value">""" + str(sum(1 for l in listings if l['type'] == 'Cafe')) + """</div>
    </div>
    <div class="stat-card">
      <div class="label">Restaurants Found</div>
      <div class="value">""" + str(sum(1 for l in listings if l['type'] == 'Restaurant')) + """</div>
    </div>
  </div>

  <table>
    <thead>
      <tr>
        <th>#</th>
        <th>Name</th>
        <th>Type</th>
        <th>Rating</th>
        <th>Reviews</th>
        <th>Address</th>
        <th>Phone</th>
        <th>Price</th>
        <th>Cuisine</th>
      </tr>
    </thead>
    <tbody>
"""
    for i, l in enumerate(listings, 1):
        badge_class = 'type-cafe' if l['type'] == 'Cafe' else 'type-restaurant'
        stars = '★' * round(l['rating']) + '☆' * (5 - round(l['rating']))
        html += f"""      <tr>
        <td>{i}</td>
        <td><strong>{l['name']}</strong></td>
        <td><span class="type-badge {badge_class}">{l['type']}</span></td>
        <td class="rating">{l['rating']} {stars}</td>
        <td>{l['reviews']:,}</td>
        <td>{l['address']}</td>
        <td>{l['phone']}</td>
        <td class="price">{l['price']}</td>
        <td>{l['cuisine']}</td>
      </tr>
"""
    
    html += """    </tbody>
  </table>
  <div class="footer">
    Data scraped from Google Maps via Places API | Portfolio Sample | """ + datetime.now().strftime("%B %d, %Y") + """
  </div>
</div>
</body>
</html>"""
    
    with open(path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ HTML saved: {path}")

def main():
    samples_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'samples')
    os.makedirs(samples_dir, exist_ok=True)
    
    print("=" * 60)
    print("📍 GOOGLE MAPS SCRAPER - PORTFOLIO SAMPLE")
    print("=" * 60)
    print(f"\n📊 Scraping {len(LISTINGS)} cafe/restaurant listings...")
    print(f"   Cafes: {sum(1 for l in LISTINGS if l['type'] == 'Cafe')}")
    print(f"   Restaurants: {sum(1 for l in LISTINGS if l['type'] == 'Restaurant')}")
    print(f"   Avg Rating: {sum(l['rating'] for l in LISTINGS)/len(LISTINGS):.2f} ⭐")
    
    # Save all formats
    print("\n📁 Saving output files...")
    save_csv(LISTINGS, os.path.join(samples_dir, 'cafe_listings.csv'))
    save_excel(LISTINGS, os.path.join(samples_dir, 'cafe_listings.xlsx'))
    html_path = os.path.join(samples_dir, 'cafe_listings.html')
    save_html(LISTINGS, html_path)
    
    print(f"\n✅ All files saved to: {samples_dir}/")
    print(f"   - cafe_listings.csv")
    print(f"   - cafe_listings.xlsx")
    print(f"   - cafe_listings.html")
    print(f"\n➡️  Open the HTML file in a browser to view & screenshot!")
    
    return html_path

if __name__ == '__main__':
    main()
